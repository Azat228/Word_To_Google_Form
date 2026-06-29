"""Export graded results into Excel workbooks.

The module creates both a summary report and a detailed per-question report,
using formatting helpers to make the spreadsheets easier to read.
"""

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
def format_submitted_time(submitted_at) -> str:
    if not submitted_at:
        return ""

    if isinstance(submitted_at, datetime):
        return submitted_at.strftime("%Y-%m-%d %H:%M:%S")

    submitted_at = str(submitted_at)

    try:
        # Google usually gives time like:
        # 2026-06-29T12:34:56.789Z
        parsed_time = datetime.fromisoformat(
            submitted_at.replace("Z", "+00:00")
        )

        return parsed_time.strftime("%Y-%m-%d %H:%M:%S")

    except ValueError:
        # fallback: manually extract time from ISO string
        if "T" in submitted_at:
            time_part = submitted_at.split("T", 1)[1]
            time_part = time_part.split(".", 1)[0]
            time_part = time_part.replace("Z", "")
            return time_part

        return submitted_at
def get_grade_fill(label: str) -> PatternFill:
    label_lower = label.lower()

    if "миним" in label_lower or "норма" in label_lower:
        return PatternFill("solid", fgColor="C6EFCE")  # green

    if "лёг" in label_lower or "лег" in label_lower:
        return PatternFill("solid", fgColor="FFEB9C")  # yellow

    if "умерен" in label_lower:
        return PatternFill("solid", fgColor="F4B183")  # orange

    if "выраж" in label_lower or "тяж" in label_lower or "высок" in label_lower:
        return PatternFill("solid", fgColor="FFC7CE")  # red

    return PatternFill("solid", fgColor="D9EAF7")


def style_header(row):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(bottom=Side(style="thin", color="808080"))

    for cell in row:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")


def auto_size_columns(ws, max_width: int = 60):
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = cell.value

            if value is None:
                continue

            max_length = max(max_length, len(str(value)))

        width = min(max_length + 2, max_width)
        ws.column_dimensions[column_letter].width = max(width, 10)


def apply_common_sheet_style(ws):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    auto_size_columns(ws)


def export_summary_report(
    graded_results: list[dict[str, Any]],
    output_path: str,
) -> str:
    # Create a compact workbook with one row per student and a summary of the
    # overall interpretation for each result.
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет"

    headers = [
        "№",
        "ФИО",
        "Класс",
        "Email",
        "Дата отправки",
        "Общий балл",
        "Макс. балл",
        "Градация",
        "Высокий риск",
        "Примечание",
    ]

    ws.append(headers)
    style_header(ws[1])

    note_text = (
        "Автоматический подсчёт. "
        "Не является диагнозом. "
        "При высоких показателях требуется проверка специалистом."
    )

    # Write one row per student and color the grade column based on the
    # resulting category label.
    for index, result in enumerate(graded_results, start=1):
        ws.append(
            [
                index,
                result.get("student_name", ""),
                result.get("student_class", ""),
                result.get("student_email", ""),
                format_submitted_time(result.get("submitted_at")),
                result.get("total_score", 0),
                result.get("max_score", 0),
                result.get("grade_label", ""),
                result.get("risk_flag", "Нет"),
                note_text,
            ]
        )

        current_row = ws.max_row
        grade_cell = ws.cell(row=current_row, column=9)
        grade_cell.fill = get_grade_fill(str(grade_cell.value))
        risk_cell = ws.cell(row=current_row, column=9)
        if risk_cell.value == "Высокий риск":
            risk_cell.fill = PatternFill("solid", fgColor="FFC7CE")
        elif risk_cell.value == "риск":
            risk_cell.fill = PatternFill("solid", fgColor="FFEB9C")
    apply_common_sheet_style(ws)

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)

    return str(output)


def export_detailed_report(
    graded_results: list[dict[str, Any]],
    output_path: str,
) -> str:
    # Create a second workbook that expands each student's result into one row
    # per question, making it easier to review the individual answers.
    wb = Workbook()
    ws = wb.active
    ws.title = "Детальный отчет"

    headers = [
        "№ ученика",
        "ФИО",
        "Класс",
        "Email",
        "Дата отправки",
        "Общий балл",
        "Макс. балл",
        "Градация",
        "Высокий риск",
        "№ вопроса",
        "Вопрос",
        "Ответ ребёнка",
        "Балл за ответ",
        "Макс. балл за вопрос",
    ]

    ws.append(headers)
    style_header(ws[1])

    student_index = 1

    for result in graded_results:
        questions = result.get("questions", [])

        sorted_questions = sorted(
            questions,
            key=lambda q: q.get("number", 0),
        )

        block_start_row = ws.max_row + 1

        for question_index, question in enumerate(sorted_questions):
            if question_index == 0:
                student_columns = [
                    student_index,
                    result.get("student_name", ""),
                    result.get("student_class", ""),
                    result.get("student_email", ""),
                    format_submitted_time(result.get("submitted_at")),
                    result.get("total_score", 0),
                    result.get("max_score", 0),
                    result.get("grade_label", ""),
                    result.get("risk_flag", "Нет"),
                ]
            else:
                student_columns = ["", "", "", "", "", "", "", "", ""]

            question_columns = [
                question.get("number", ""),
                question.get("title", ""),
                question.get("answer", ""),
                question.get("score", 0),
                question.get("max_score", 0),
            ]

            ws.append(student_columns + question_columns)

        block_end_row = ws.max_row

        # Merge student-level cells vertically.
        # Columns A-H belong to the student, not to every individual question.
        if block_end_row > block_start_row:
            for column in range(1, 10):
                ws.merge_cells(
                    start_row=block_start_row,
                    start_column=column,
                    end_row=block_end_row,
                    end_column=column,
                )

                merged_cell = ws.cell(row=block_start_row, column=column)
                merged_cell.alignment = Alignment(
                    vertical="center",
                    horizontal="center",
                    wrap_text=True,
                )

        # Color the merged grade cell.
        grade_cell = ws.cell(row=block_start_row, column=9)
        grade_cell.fill = get_grade_fill(str(grade_cell.value))
        risk_cell = ws.cell(row=block_start_row, column=9)

        if risk_cell.value == "Высокий риск":
            risk_cell.fill = PatternFill("solid", fgColor="FFC7CE")
        elif risk_cell.value == "риск":
            risk_cell.fill = PatternFill("solid", fgColor="FFEB9C")

        student_index += 1

    ws.freeze_panes = "A2"

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Re-apply centered alignment for merged student info cells.
    for merged_range in ws.merged_cells.ranges:
        top_left_cell = ws.cell(
            row=merged_range.min_row,
            column=merged_range.min_col,
        )
        top_left_cell.alignment = Alignment(
            vertical="center",
            horizontal="center",
            wrap_text=True,
        )

    auto_size_columns(ws)

    # Make question/answer columns wider.
    ws.column_dimensions["K"].width = 35  # Вопрос
    ws.column_dimensions["L"].width = 60  # Ответ ребёнка

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)

    return str(output)

def export_reports(
    graded_results: list[dict[str, Any]],
    output_dir: str,
) -> tuple[str, str]:
    # Generate both the summary and detailed Excel files from the same graded
    # dataset so the user can inspect the results at different levels of detail.
    output_folder = Path(output_dir)
    output_folder.mkdir(parents=True, exist_ok=True)

    summary_path = output_folder / "отчет.xlsx"
    detailed_path = output_folder / "детальный_отчет.xlsx"

    export_summary_report(
        graded_results=graded_results,
        output_path=str(summary_path),
    )

    export_detailed_report(
        graded_results=graded_results,
        output_path=str(detailed_path),
    )

    return str(summary_path), str(detailed_path)